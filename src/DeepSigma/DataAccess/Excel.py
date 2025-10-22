from openpyxl import load_workbook
import xlrd
import pandas as pd
from openpyxl.workbook import Workbook

class Excel:
    def __init__(self, full_file_path: str) -> None:
        self.__full_file_path: str = full_file_path
        self.__is_xls_file_format :bool = self.is_xls_file()

    def is_xls_file(self) -> bool:
        if self.__full_file_path.endswith('.xls'):
            return True
        return False

    def get_cell_value(self, sheet_name: str, cell_name: str):
        """
        Gets value from a specified cell in an Excel worksheet.

        Usage example: Excel.get_cell_value('Sheet1', 'A1')
        :param sheet_name:
        :param cell_name:
        :return:
        """
        if self.__is_xls_file_format:
            workbook: xlrd.Book = xlrd.open_workbook(self.__full_file_path)
            sheet = workbook.sheet_by_name(sheet_name)
            row_index, column_index = Excel.__cell_to_indices(cell_name)
            value = sheet.cell_value(row_index, column_index)
            return value
        else:
            workbook: Workbook = load_workbook(self.__full_file_path)
            sheet = workbook[sheet_name]
            return sheet[cell_name].value

    def get_column_values(self, sheet_name: str, column_letter_index: str, start_row_index: int = 0):
        """
        Gets values from a specified column in an Excel worksheet.

        Usage example: Excel.get_column_values('Sheet1', 'A')
        :param sheet_name:
        :param column_letter_index:
        :param start_row_index:
        :return:
        """
        column_values: list = []
        if self.__is_xls_file_format:
            workbook: xlrd.Book = xlrd.open_workbook(self.__full_file_path)
            sheet = workbook.sheet_by_name(sheet_name)
            column_values = sheet.col_values(column_letter_index, start_rowx=start_row_index)
        else:
            workbook: Workbook = load_workbook(self.__full_file_path)
            sheet = workbook[sheet_name]
            all_column_cells = [cell.value for cell in sheet[column_letter_index]]
            column_index: int = Excel.__cell_to_indices(column_letter_index+"1")[1]
            column_values = [value for value in all_column_cells[start_row_index:]]
        return column_values

    def get_row_values(self, sheet_name: str, row_index: int, start_column_index: int = 0):
        """
        Gets values from a specified row in an Excel worksheet.

        Usage example: Excel.get_row_values('Sheet1', 1)
        :param sheet_name:
        :param row_index:
        :param start_column_index:
        :return:
        """
        row_values: list = []
        if self.__is_xls_file_format:
            workbook: xlrd.Book = xlrd.open_workbook(self.__full_file_path)
            sheet = workbook.sheet_by_name(sheet_name)
            row_values = sheet.row_values(row_index, start_colx=start_column_index)
        else:
            workbook: Workbook = load_workbook(self.__full_file_path)
            sheet = workbook[sheet_name]
            all_row_cells = [cell.value for cell in sheet[row_index]]
            row_values = [value for value in all_row_cells[start_column_index:]]
        return row_values

    def get_data_as_datatable(self, sheet_name: str, header_index: int = 0,
                              column_to_designate_as_index: int = None) -> pd.DataFrame:
        """
        Gets data from an Excel worksheet as a Pandas data frame.
        :param sheet_name:
        :param header_index:
        :param column_to_designate_as_index:
        :return:
        """
        data_frame = pd.read_excel(self.__full_file_path, sheet_name=sheet_name, header=header_index,
                                   index_col=column_to_designate_as_index)
        return data_frame


    @staticmethod
    def __cell_to_indices(cell_ref) -> (int, int):
        """Convert Excel-style cell reference (e.g. 'B2') to zero-based (row, col)."""
        col_str = ''.join(filter(str.isalpha, cell_ref)).upper()
        row_str = ''.join(filter(str.isdigit, cell_ref))

        col = 0
        for char in col_str:
            col = col * 26 + (ord(char) - ord('A') + 1)
        return int(row_str) - 1, col - 1  # zero-based row and col

